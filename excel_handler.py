import openpyxl
import database as dbtools
import decimal
import os
from datetime import datetime


def export_table_as_excel(table_name, table_column=None, compare=None):
    # Connection to Database
    conn = dbtools.connect_to_db(dbtools.config_file)
    cursor = dbtools.create_cursor(conn)

    if table_column is None and compare is None:
        # SQL-Abfrage, um die gesamte Tabelle abzurufen
        query = f"SELECT * FROM {table_name}"
    elif table_column is not None and compare is not None:
        # SQL-Abfrage, um den gesuchten Mitarbeiter zu finden
        query = f"SELECT * FROM {table_name} WHERE {table_column} = '{compare}'"

    # Query ausführen
    cursor.execute(query)

    # Spaltennamen (Überschriften) abrufen
    column_names = [desc[0] for desc in cursor.description]

    # Daten abrufen
    data = []

    for row in cursor:
        # Konvertiere Decimal-Werte in float
        converted_row = [float(value) if isinstance(value, decimal.Decimal) else value for value in row]
        data.append(converted_row)

    # Excel-Datei erstellen
    wb = openpyxl.Workbook()
    ws = wb.active

    # Schreibe Spaltennamen (Überschriften) in die erste Zeile
    for col_idx, column_name in enumerate(column_names, start=1):
        ws.cell(row=1, column=col_idx, value=column_name)

    # Schreibe Daten in Excel-Tabelle
    for row_idx, row_data in enumerate(data, start=2):  # Beginne bei Zeile 2 für Daten
        for col_idx, cell_data in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=cell_data)

    # Ordner für den Export erstellen (mit aktuellem Datum und Uhrzeit)
    export_folder = datetime.now().strftime("%Y_%m_%d-%H_%M")
    os.makedirs(f"output/{export_folder}", exist_ok=True)

    # Excel-Datei im exportierten Ordner speichern
    excel_filename = f"output/{export_folder}/{table_name}.xlsx"
    wb.save(excel_filename)
    wb.close()

    print(f"{table_name} wurde als Excel-Datei exportiert in {export_folder}.")
    return True